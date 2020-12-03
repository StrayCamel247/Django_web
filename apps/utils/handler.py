#!/usr/bin/python
# -*- coding: utf-8 -*-
# __author__ : stray_camel
# __date__: 2020/05/26 11:09:55

import io
import socket
import re
import requests
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, PatternFill, Side
from openpyxl.utils import get_column_letter


def get_host_ip():
    """查询本机内网ip地址1"""
    try:
        s = socket.socket(socket.AF_INET, socket.SOCK_DGRAM)
        s.connect(('8.8.8.8', 80))
        ip = s.getsockname()[0]
    finally:
        s.close()
    return ip

def get_client_ip(request):
    """查询本机内网ip地址2"""
    ip_type = 'Client_Real_IP'
    ip = request.headers.get(ip_type, '')
    if ip == '':
        ip_type = 'X_Real_IP'
        ip = request.headers.get(ip_type, '')
    if ip == '':
        ip_type = 'X-Forwarded-For'
        ip = request.headers.get(ip_type, '')
    if ip == '':
        ip_type = 'remote_addr'
        ip = getattr(request, ip_type, '')
    if ip == '':
        ip_type = ''
    return ip, ip_type
def get_internet_ip():
    """获取本机的外网地址"""
    res = requests.get('http://members.3322.org/dyndns/getip').content.decode('utf-8')
    ip = re.match(r"\d+\.\d+\.\d+\.\d*",res).group()
    return ip

def get_local_host_ip(params=None):
    """获取本机"""
    # 外网ip
    ip = get_internet_ip()
    # 太网适配器 IPV4:
    # ip = socket.gethostbyname(socket.gethostname())
    # 局域网 IPV4
    ip = get_host_ip()
    return ip


def style_range(ws, cell_range=None, border=None, fill=None, alignment=None, merge_header=False,
                auto_width=False, merge_list=None, fill_row_list=None, comment_msg=None):
    """
    # excel设置范围样式函数
    :param ws: 表格对象
    :param cell_range:  图表类型
    :param border: 边框样式
    :param fill: 颜色填充样式
    :param alignment:  对齐样式
    :param merge_header： 表头是否合并并居中
    :param auto_width：是否根据内容自动设置列宽
    :param merge_list 要合并的行列数据
    :param fill_row_list 要填充颜色的行数据
    :param comment_msg 添加标注的字典数据
    :return:
    """
    if border is None:
        # 定义边框填充
        bd = Side(border_style='thin', color='000000')
        border = Border(left=bd, right=bd, top=bd, bottom=bd)
    if alignment is None:
        alignment = Alignment(horizontal='center', vertical='center')

    if cell_range == 'product_target':
        # 设置前四行合并数据
        ws.merge_cells(start_row=1, end_row=1, start_column=1,
                       end_column=ws.max_column)
        if current_user.role_name == "管理员":
            ws.merge_cells(start_row=2, end_row=2,
                           start_column=6, end_column=ws.max_column)
            ws.merge_cells(start_row=2, end_row=3,
                           start_column=1, end_column=1)
            ws.merge_cells(start_row=2, end_row=3,
                           start_column=2, end_column=2)
            ws.merge_cells(start_row=2, end_row=3,
                           start_column=3, end_column=3)
            ws.merge_cells(start_row=2, end_row=3,
                           start_column=4, end_column=4)
            ws.merge_cells(start_row=2, end_row=3,
                           start_column=5, end_column=5)
            ws.merge_cells(start_row=4, end_row=4,
                           start_column=1, end_column=3)
            if merge_list:
                for i in merge_list:
                    ws.merge_cells(start_row=i.get("start_row"), end_row=i.get("end_row"),
                                   start_column=i.get("start_column"), end_column=i.get("end_column"))
        else:
            if merge_list:
                for i in merge_list:
                    ws.merge_cells(start_row=i.get("start_row"), end_row=i.get("end_row"),
                                   start_column=i.get("start_column"), end_column=i.get("end_column"))
        # import pdb;pdb.set_trace()
        # 第一行行高
        ws.row_dimensions[1].height = 40
        ws['A1'].alignment = alignment
        # 设置每一行的高
        for row in range(2, ws.max_row + 1):
            # 其他行行高
            ws.row_dimensions[row].height = 25
        # A2：E20 第一列字母+起始行数 :  最后一列字母  + 最后一行
        #                A   2    :   E           20
        cell_range = get_column_letter(1) + "2:" + get_column_letter(ws.max_column) + str(
            ws.max_row)

        rows = ws[cell_range]

        column_widths = dict()
        # 填充颜色
        blue = PatternFill(start_color='87CEFA',
                           end_color='87CEFA', fill_type='solid')  # 蓝色
        orange = PatternFill(start_color='FF7F50',
                             end_color='FF7F50', fill_type='solid')  # 橙色
        # # 设置每一列的格式（居中）无效
        # for column in range(1, ws.max_column + 1):
        #     # print(get_column_letter(column))
        #     ws.column_dimensions[get_column_letter(column)].alignment = alignment
        # # 设置每一行的格式（颜色填充）无效
        # for row in range(1, ws.max_row + 1):
        #     if row == 4:
        #         ws.row_dimensions[row].fill = orange
        #     elif row in fill_row_list:
        #         ws.row_dimensions[row].fill = blue
        if current_user.role_name == "管理员":
            for row in rows:
                for c in row:
                    # 第四行填充
                    if c.row == 4:
                        c.fill = orange
                    if c.row in fill_row_list:
                        c.fill = blue
                    if alignment:
                        c.alignment = alignment
                    if border:
                        c.border = border
        else:
            for row in rows:
                for c in row:
                    # 第四行填充
                    if c.row == 2 or c.row == 3:
                        c.fill = orange
                    if alignment:
                        c.alignment = alignment
                    if border:
                        c.border = border
        # ws.cell(row=4, column=1, value='合计').fill = orange
        # for i in fill_row_list:
        #     ws.cell(row=i.get("minor_total")).fill = blue
        # 列宽自适应
        if auto_width:
            for row in rows:
                for c in row:
                    if current_user.role_name == "管理员":
                        if c.row == 4:
                            c.fill = orange
                        if c.row in fill_row_list:
                            c.fill = blue
                        if alignment:
                            c.alignment = alignment
                        if border:
                            c.border = border
                    else:
                        # 第四行填充
                        if c.row == 2 or c.row == 3:
                            c.fill = orange
                        if alignment:
                            c.alignment = alignment
                        if border:
                            c.border = border
                    if c.value is not None:
                        if isinstance(c.value, datetime.datetime) or isinstance(c.value, datetime.date):
                            cell_width = len(str(c.value))
                        elif isinstance(c.value, str):
                            cell_width = len(c.value.encode('utf-8'))
                        else:
                            cell_width = len(str(c.value))
                        if c.column not in column_widths:
                            column_widths.update({c.column: cell_width})
                        else:
                            if cell_width > column_widths.get(c.column, 20):
                                # c.column列（A,B,C,D），cell_width宽度
                                column_widths.update({c.column: cell_width})

            # 设置列宽
            for k, v in column_widths.items():
                ws.column_dimensions[k].width = v
    elif cell_range == 'product_sale':
        # 设置前四行合并数据
        ws.merge_cells(start_row=1, end_row=1, start_column=1,
                       end_column=ws.max_column)
        if current_user.role_name == "管理员":
            if merge_list:
                for i in merge_list:
                    ws.merge_cells(start_row=i.get("start_row"), end_row=i.get("end_row"),
                                   start_column=i.get("start_column"), end_column=i.get("end_column"))
            # 保证最后一个类别合并
            if merge_list:
                if merge_list[-1] == {'start_row': 2, 'end_row': 2, 'start_column': 17, 'end_column': 18}:
                    ws.merge_cells(start_row=4, end_row=ws.max_row,
                                   start_column=1, end_column=1)
                else:
                    ws.merge_cells(start_row=merge_list[-1].get("end_row") + 1, end_row=ws.max_row,
                                   start_column=1, end_column=1)
            else:
                ws.merge_cells(start_row=3, end_row=ws.max_row,
                               start_column=1, end_column=1)
        else:
            if merge_list:
                for i in merge_list:
                    ws.merge_cells(start_row=i.get("start_row"), end_row=i.get("end_row"),
                                   start_column=i.get("start_column"), end_column=i.get("end_column"))
        # 第一行行高
        ws.row_dimensions[1].height = 40
        ws['A1'].alignment = alignment
        # 设置每一行的高
        for row in range(2, ws.max_row + 1):
            # 其他行行高
            ws.row_dimensions[row].height = 25
        # A2：E20 第一列字母+起始行数 :  最后一列字母  + 最后一行
        #                A   2    :   E           20
        cell_range = get_column_letter(1) + "2:" + get_column_letter(ws.max_column) + str(
            ws.max_row)
        rows = ws[cell_range]

        column_widths = dict()

        # 列宽自适应
        if auto_width:
            for row in rows:
                for c in row:
                    # 设置每一个单元格的填充颜色，对齐样式，边框
                    # 第四行填充
                    if fill:
                        c.fill = fill
                    if alignment:
                        c.alignment = alignment
                    if border:
                        c.border = border
                    if c.value is not None:
                        if isinstance(c.value, datetime.datetime) or isinstance(c.value, datetime.date):
                            cell_width = len(str(c.value))
                        elif isinstance(c.value, str):
                            cell_width = len(c.value.encode('utf-8'))
                        else:
                            cell_width = len(str(c.value))

                        if c.column not in column_widths:
                            column_widths.update({c.column: cell_width})
                        # 当前单元格数据大于上一次属于同一列数据， 取大的一列数据宽度
                        else:
                            if cell_width > column_widths.get(c.column, 20):
                                # c.column列（A,B,C,D），cell_width宽度
                                column_widths.update({c.column: cell_width})
                                # print(column_widths)

            # 设置列宽
            for k, v in column_widths.items():
                ws.column_dimensions[k].width = v
    else:
        # 合并第一行单元格
        if merge_header:
            ws.merge_cells(start_row=1, end_row=1,
                           start_column=1, end_column=ws.max_column)
            # 设置第一行行高
            ws.row_dimensions[1].height = 40
            # 设置水平垂直居中
            ws['A1'].alignment = alignment
            # 设置每一行的高
            for row in range(2, ws.max_row + 1):
                ws.row_dimensions[row].height = 25
            # 表格起始单元格和结束单元格 A2:K12
            cell_range = get_column_letter(1) + "2:" + get_column_letter(ws.max_column) + str(
                ws.max_row)
        else:
            cell_range = ws.dimensions
            if comment_msg:
                for i in range(1, len(comment_msg)+1):
                    comment = Comment(comment_msg.get(i), 'owner')
                    ws['%s1' % get_column_letter(i)].comment = comment

        if merge_list:
            for i in merge_list:
                ws.merge_cells(start_row=i.get("start_row"), end_row=i.get("end_row"),
                               start_column=i.get("start_column"), end_column=i.get("end_column"))

        rows = ws[cell_range]
        # 列宽字典
        column_widths = dict()

        # 列宽自适应
        if auto_width:
            for row in rows:
                for c in row:
                    # 设置每一个单元格的填充颜色，对齐样式，边框
                    # 第四行填充
                    if fill:
                        c.fill = fill
                    if alignment:
                        c.alignment = alignment
                    if border:
                        c.border = border
                    if c.value is not None:
                        # 获取每个单元格的数据大小从而设置宽度
                        if isinstance(c.value, datetime.datetime) or isinstance(c.value, datetime.date):
                            cell_width = len(str(c.value))
                        elif isinstance(c.value, str):
                            cell_width = len(c.value.encode('utf-8'))
                        else:
                            cell_width = len(str(c.value))
                        # 设置单元格的宽度，同一列的单元格大的宽度覆盖小的宽度
                        if c.column not in column_widths:
                            column_widths.update({c.column: cell_width})
                        else:
                            if cell_width > column_widths.get(c.column, 20):
                                # c.column列（A,B,C,D），cell_width宽度
                                column_widths.update({c.column: cell_width})

            # # 设置列宽
            # for k, v in column_widths.items():
            #     ws.column_dimensions[k].width = v + 1


def list_to_excel_file(file_name, data_dict, header_name=''):
    wb = Workbook()
    ws = wb.active

    for i in data_dict:
        ws.append(i)

    style_range(ws, merge_header=False, auto_width=True)
    ws.freeze_panes = ws['A2']

    # 新建一个string IO流
    output = io.BytesIO()
    # excel写入到IO流中
    wb.save(output)
    # output.seek(0)
    #
    # return send_file(output, mimetype='application/vnd.ms-excel', cache_timeout=0,
    #           attachment_filename=file_name.encode().decode('latin-1'), as_attachment=True)

    rv = make_response(output.getvalue())
    # 特别注意，必须加这一行
    output.close()
    # mime_type = mimetypes.guess_type('%s.xlsx' % file_name.encode().decode('latin-1'))[0]
    rv.headers['Content-Type'] = "application/vnd.ms-excel"
    rv.headers["Cache-Control"] = "no-cache"
    # rv.headers['Content-Disposition'] = 'attachment; filename={}.xlsx'.format(file_name.encode().decode('latin-1'))
    # rv.headers['Content-Disposition'] = 'attachment; filename=replenish.xlsx'.format(file_name.encode().decode('latin-1'))
    rv.headers['Content-Disposition'] = 'attachment; filename={file_name}'.format(
        file_name=file_name.encode().decode('latin-1') if file_name else 'replenish.xlsx')

    return rv
